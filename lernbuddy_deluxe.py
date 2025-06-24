import streamlit as st 
import datetime
import pandas as pd
import json
import os
import requests
from streamlit_lottie import st_lottie
from fpdf import FPDF
from ics import Calendar, Event
from openai import OpenAI

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

PRIMARY = "#003865"
SECONDARY = "#00A3E0"
ACCENT = "#F39200"

if "darkmode" not in st.session_state:
    st.session_state.darkmode = False

def toggle_darkmode():
    st.session_state.darkmode = not st.session_state.darkmode

st.set_page_config(page_title="Lernbuddy Deluxe", layout="wide")

st.markdown(f"""
    <style>
    body {{
        background-color: {"#1e1e1e" if st.session_state.darkmode else "#ffffff"};
        color: {"#ffffff" if st.session_state.darkmode else "#000000"};
    }}
    .stButton>button {{
        background-color: {ACCENT};
        color: white;
        border-radius: 5px;
    }}
    </style>
""", unsafe_allow_html=True)

def show_lottie(url, h=200):
    r = requests.get(url)
    if r.status_code == 200:
        st_lottie(r.json(), height=h)
    else:
        st.warning("⚠️ Animation konnte nicht geladen werden.")

with st.sidebar:
    st.title("📚 Lernbuddy Deluxe")
    st.button("🌗 Darkmode umschalten", on_click=toggle_darkmode)
    menu = st.radio("Navigation", ["🏠 Start", "💬 GPT-Chat", "🧠 Lernplan", "🔎 Suche", "🎓 Hochschule"])

# Startseite
if menu == "🏠 Start":
    st.title("🎓 Willkommen bei Lernbuddy Deluxe 👋")
    show_lottie("https://assets2.lottiefiles.com/packages/lf20_myejiggj.json")
    st.markdown("""

**Lernbuddy Deluxe** ist **mehr als nur ein Chatbot** – er ist dein persönlicher Studien-Coach, digitaler Lernpartner und smarter Assistent, der dich durch das gesamte Semester begleitet! 🚀📚

---

## 💡 Was Lernbuddy Deluxe für dich tun kann:

### 💬 GPT-Chat – Dein KI-Tutor  
Stelle Fragen rund ums Studium – oder auch zum Leben. Ob:  
- ✅ Lernhilfe & Verständnisfragen  
- ✅ Zusammenfassungen & Erklärungen  
- ✅ Studienorganisation oder Alltagssorgen  
**Der GPT-Tutor ist für dich da!**

### 🧠 Automatischer Lernplan-Generator  
- Du trägst deine Fächer, Prüfungen & Schwierigkeitsgrad ein  
- Der Bot erstellt dir automatisch einen effizienten Lernplan – taggenau mit Zeitvorgaben  
- **Exportiere den Plan als PDF oder .ics-Kalenderdatei**

### 🔎 Intelligente Suchfunktion  
Finde blitzschnell Inhalte und Fächer im Lernplan wieder – perfekt zum Wiederholen!

### 🎨 Farben & Darkmode  
Wähle deinen Style:  
- 🌗 Darkmode  
- 🎨 5 moderne Farbpaletten  
- Inspiriert vom Design der **Hochschule Kempten**

---

## ✨ Entwickelt für Studierende – von Studierenden  
> Mit ❤️ von **Taner Altin** & **Shefki Kuleta**  
> Powered by **Streamlit** & **OpenAI GPT-4**
""")

# GPT-Chat
elif menu == "💬 GPT-Chat":
    st.header("💬 GPT-Chat")
    user_color = st.color_picker("Farbe für deine Nachrichten", "#00A3E0")
    bot_color = st.color_picker("Farbe für GPT", "#F39200")

    if "chat" not in st.session_state:
        st.session_state.chat = []

    user_input = st.text_input("🗨️ Deine Frage:")
    if user_input:
        st.session_state.chat.append({"role": "user", "content": user_input})
        try:
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=st.session_state.chat
            )
            reply = response.choices[0].message.content
            st.session_state.chat.append({"role": "assistant", "content": reply})
        except Exception as e:
            reply = f"Fehler: {e}"
            st.session_state.chat.append({"role": "assistant", "content": reply})

    for msg in st.session_state.chat[::-1]:
        color = user_color if msg["role"] == "user" else bot_color
        sender = "👤 Du" if msg["role"] == "user" else "🤖 GPT"
        st.markdown(f"""
        <div style='background-color:{color};padding:10px;border-radius:10px;margin-bottom:10px;color:white'>
        <strong>{sender}:</strong><br>{msg['content']}
        </div>
        """, unsafe_allow_html=True)

# Lernplan
elif menu == "🧠 Lernplan":
    st.header("🧠 Lernplan mit GPT-Hinweisen, Farben, Excel- & Kalender-Export + Statistik")

    # 1) Eingabe: Fächer, Prüfungstermine, Schwierigkeit, GPT-Hinweise
    start_date = st.date_input("📆 Startdatum des Lernplans", datetime.date.today())
    n = st.number_input("Wie viele Prüfungen hast du?", 1, 10)
    subjects = []
    for i in range(int(n)):
        name = st.text_input(f"📘 Fach {i+1}", key=f"subj_{i}")
        date_input = st.date_input(f"📅 Prüfung {i+1}", key=f"date_{i}")
        diff = st.slider("📊 Schwierigkeit (1–10)", 1, 10, key=f"diff_{i}")
        hint = st.text_area(
            f"🧠 Hinweis für GPT zu '{name or 'Fach'}'",
            key=f"hint_{i}",
            placeholder="z. B.: nur ab 12 Uhr, nicht sonntags …"
        )
        if name.strip():
            subjects.append({
                "name":       name.strip(),
                "exam_date":  str(date_input),
                "difficulty": diff,
                "hint":       hint.strip() or "keine"
            })

    # 2) Button zum GPT-Plan
    if st.button("✅ GPT-Lernplan generieren"):
        if not subjects:
            st.warning("⚠️ Bitte gib mindestens ein Fach ein.")
        else:
            fachliste = "\n".join(
                f"- {s['name']} (Prüfung: {s['exam_date']}, Schwierigkeit: {s['difficulty']}) – Hinweis: {s['hint']}"
                for s in subjects
            )
            prompt = f""" 
Du bist ein Lerncoach und erstellst einen Lernplan für diese Fächer, Prüfungen und individuellen Hinweise.
Der Plan startet ab dem {{start_date}}.

Erstelle einen 4-Wochen-Plan mit Uhrzeiten (z. B. 10:00–10:45), Pausen und max. 4 Blöcken pro Tag.
Beachte persönliche Wünsche und prüfe auf Überschneidungen.

Fächer & Hinweise:
{fachliste}

Gib den Plan im Format aus:

Montag, 01.07.2025
- 12:00–12:45: Mathe
- 14:00–14:45: BWL
"""
            with st.spinner("GPT plant deinen Lernplan …"):
                try:
                    # 3) GPT-Aufruf
                    response = client.chat.completions.create(
                        model="gpt-4",
                        messages=[{"role": "user", "content": prompt}]
                    )
                    result = response.choices[0].message.content
                    st.markdown("### 📅 Vorschlag von GPT:")
                    st.markdown(result)

                    # 4) Parser für verschiedene GPT-Formate
                    import re, pandas as pd
                    def parse_gpt_plan(text):
                        date_re = re.compile(r"^([A-Za-zäöüÄÖÜß]+)[, ]+\s*(\d{2}[./]\d{2}[./]\d{4})")
                        session_re = re.compile(
                            r"-?\s*(\d{2}[:\.]\d{2})\s*(?:–|-|bis)\s*(\d{2}[:\.]\d{2})\s*[:\-–]?\s*(.+)",
                            re.IGNORECASE
                        )
                        rows = []
                        current = {}
                        for raw in text.splitlines():
                            line = raw.strip()
                            if not line or "pause" in line.lower():
                                continue
                            dm = date_re.match(line)
                            sm = session_re.match(line)
                            if dm:
                                current = {
                                    "Wochentag": dm.group(1),
                                    "Datum":     dm.group(2)
                                }
                            elif sm and current:
                                start, end, fach = sm.groups()
                                rows.append({
                                    "Wochentag": current["Wochentag"],
                                    "Datum":     current["Datum"],
                                    "Fach":      fach.strip(),
                                    "Startzeit": start.replace(".", ":"),
                                    "Endzeit":   end.replace(".", ":"),
                                    "Dauer":     "45 Min"
                                })
                        return pd.DataFrame(rows)

                    df_gpt = parse_gpt_plan(result)

                    if df_gpt.empty:
                        st.warning("⚠️ GPT-Plan konnte nicht in eine Tabelle umgewandelt werden.")
                    else:
                        # 5) Farbige Anzeige in Streamlit
                        st.markdown("### 🧠 GPT-Plan (farbig):")
                        palette = ["#FFD700", "#00CED1", "#FF8C00", "#ADFF2F",
                                   "#DA70D6", "#FFA07A", "#7FFFD4", "#D2691E"]
                        fachfarben = {f: palette[i % len(palette)] for i, f in enumerate(df_gpt["Fach"].unique())}

                        for _, r in df_gpt.iterrows():
                            c = fachfarben.get(r["Fach"], "#EEE")
                            st.markdown(f"""
                                <div style='background-color:{c};padding:8px;border-radius:6px;margin-bottom:6px'>
                                  <strong>{r["Datum"]} ({r["Wochentag"]})</strong><br>
                                  {r["Startzeit"]}–{r["Endzeit"]}: <b>{r["Fach"]}</b>
                                </div>
                            """, unsafe_allow_html=True)

                        # 6) 💡 Motivation & Tipps
                        st.markdown("### 💡 Motivation & Tipps")
                        st.markdown(
                            "- Pausen stärken den Fokus (5 Min Dehnen oder Atmen)\n"
                            "- Wiederholung ist King – kurz vorm Schlaf nochmal überfliegen\n"
                            "- Schlaf & Bewegung helfen dem Gehirn, neues Wissen zu verankern"
                        )

                        # 7) Lernzeit-Statistik
                        df_stats = df_gpt.copy()
                        df_stats["Minuten"] = df_stats["Dauer"].str.extract(r"(\d+)").astype(int)
                        stats = df_stats.groupby("Fach").agg(
                            Sessions=("Fach", "count"),
                            Total_Minuten=("Minuten", "sum")
                        ).reset_index()
                        st.markdown("### 📊 Lernzeit-Statistik pro Fach")
                        st.dataframe(stats)

                        # 8) Excel-Export mit Farbformatierung
                        import openpyxl
                        from openpyxl.styles import Font, PatternFill
                        from openpyxl.utils.dataframe import dataframe_to_rows

                        def export_excel(df, stats_df, filename="lernplan.xlsx"):
                            wb = openpyxl.Workbook()
                            ws1 = wb.active; ws1.title = "Plan"
                            for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                                ws1.append(row)
                                for j, cell in enumerate(ws1[i], 1):
                                    if i == 1:
                                        cell.font = Font(bold=True)
                                    elif df.columns[j-1] == "Fach":
                                        col = fachfarben[cell.value].lstrip("#")
                                        cell.fill = PatternFill(start_color=col, end_color=col, fill_type="solid")
                            ws2 = wb.create_sheet("Statistik")
                            for i, row in enumerate(dataframe_to_rows(stats_df, index=False, header=True), 1):
                                ws2.append(row)
                                if i == 1:
                                    for cell in ws2[i]:
                                        cell.font = Font(bold=True)
                            for ws in (ws1, ws2):
                                for col in ws.columns:
                                    width = max(len(str(c.value)) for c in col) + 2
                                    ws.column_dimensions[col[0].column_letter].width = width
                            wb.save(filename)
                            with open(filename, "rb") as f:
                                st.download_button("📥 Excel herunterladen", f, file_name=filename)

                        export_excel(df_gpt, stats)

                        # 9) ICS-Kalender-Export
                        from ics import Calendar, Event
                        import datetime as dt

                        def export_ics(df, filename="lernplan.ics"):
                            cal = Calendar()
                            for _, r in df.iterrows():
                                d = dt.datetime.strptime(r["Datum"], "%d.%m.%Y").date()
                                stime = dt.datetime.strptime(r["Startzeit"], "%H:%M").time()
                                etime = dt.datetime.strptime(r["Endzeit"], "%H:%M").time()
                                ev = Event()
                                ev.name        = f"{r['Fach']} – Lernen"
                                ev.begin       = dt.datetime.combine(d, stime)
                                ev.end         = dt.datetime.combine(d, etime)
                                ev.description = f"{r['Fach']} am {r['Datum']} ({r['Wochentag']})"
                                cal.events.add(ev)
                            with open(filename, "w", encoding="utf-8") as f:
                                f.writelines(cal)
                            with open(filename, "rb") as f:
                                st.download_button(
                                    "📆 ICS herunterladen",
                                    f,
                                    file_name=filename,
                                    mime="text/calendar"
                                )

                        export_ics(df_gpt)

                except Exception as e:
                    st.error(f"Fehler beim GPT-Aufruf: {e}")




# Suche
elif menu == "🔎 Suche":
    st.header("🔎 Lernplan durchsuchen")
    term = st.text_input("🔍 Suchbegriff:")
    if os.path.exists("lernplan.xlsx"):
        df = pd.read_excel("lernplan.xlsx")
        result = df[df["Fach"].str.contains(term, case=False, na=False)]
        if not result.empty:
            st.success(f"{len(result)} Einträge gefunden:")
            st.dataframe(result)
        else:
            st.warning("Keine Treffer.")
    else:
        st.info("Noch kein Lernplan vorhanden.")

# Hochschule
elif menu == "🎓 Hochschule":
    def load_lottie_url(url):
        try:
            r = requests.get(url)
            if r.status_code != 200:
                return None
            return r.json()
        except Exception:
            return None

    # --- Header & Animation ---
    st.markdown(
        """
        <div style="display:flex; align-items:center; gap:1rem;">
          <img src="https://www.hs-kempten.de/fileadmin/favicon/android-chrome-192x192.png"
               style="width:60px; border-radius:12px;">
          <h1 style="margin:0; font-size:2.2rem;">🎓 Hochschule Kempten</h1>
        </div>
        """,
        unsafe_allow_html=True
    )

    lottie_json = load_lottie_url("https://assets10.lottiefiles.com/packages/lf20_3rwasyjy.json")
    if lottie_json:
        st_lottie(lottie_json, height=200)
    else:
        st.warning("📌 Animation konnte nicht geladen werden.")

    c1, c2, c3 = st.columns(3)
    c1.metric("📚 Studiengänge", "40+")
    c2.metric("👩‍🎓 Studierende", "5.000+")
    c3.metric("🏫 Fakultäten", "5")

    tabs = st.tabs(["🔗 Links", "🍽️ Mensaplan", "📖 Bibliothek", "💻 Moodle", "🗺️ Campus-Karte"])

    # Tab 1: Links
    with tabs[0]:
        st.subheader("🔗 Wichtige Links")
        cards = [
            ("🌐 Website", "https://www.hs-kempten.de/"),
            ("📚 Studiengänge", "https://www.hs-kempten.de/studium/studiengaenge"),
            ("🍽️ Mensaplan", "https://www.hs-kempten.de/campusgastronomie"),
            ("📖 Bibliothek", "https://www.hs-kempten.de/bibliothek"),
            ("💻 Moodle", "https://moodle.hs-kempten.de/login/"),
            ("🧾 MeinCampus", "https://meincampus.hs-kempten.de/qisserver/pages/cs/sys/portal/hisinoneStartPage.faces")
        ]
        cols = st.columns(3)
        for i, (label, url) in enumerate(cards):
            with cols[i % 3]:
                st.markdown(f"""
                <div style="
                    border:1px solid #ddd;
                    border-radius:8px;
                    padding:1rem;
                    text-align:center;
                    box-shadow:2px 2px 6px rgba(0,0,0,0.1);
                    transition:transform .2s;
                  "
                  onmouseover="this.style.transform='scale(1.03)';"
                  onmouseout="this.style.transform='scale(1)';"
                >
                  <h4 style="margin-bottom:0.5rem;">{label}</h4>
                  <a href="{url}" target="_blank" style="color:#00CED1;">{url}</a>
                </div>
                """, unsafe_allow_html=True)

    # Tab 2: Mensaplan
    with tabs[1]:
        st.subheader("🍽️ Mensaplan der Woche")

        pdf_url = "https://www.max-manager.de/daten-extern/augsburg/pdf/wochenplaene/hs-kempten/aktuell.pdf"

        try:
            st.markdown(f"""
            👉 Den aktuellen Mensaplan findest du hier als PDF:  
            [📄 Jetzt anzeigen]({pdf_url})
            """)
            st.download_button(
                label="📥 PDF herunterladen",
                data=requests.get(pdf_url).content,
                file_name="mensaplan_kempten.pdf",
                mime="application/pdf"
            )
        except Exception as e:
            st.error(f"❌ Fehler beim Laden des Mensaplan-PDFs: {e}")

    # Tab 3: Bibliothek
    with tabs[2]:
        with st.expander("📖 Öffnungszeiten & Services"):
            st.markdown("""
            - Mo–Fr: 08:00–20:00  
            - Sa: 10:00–16:00  
            - Buchkatalog: [online suchen](https://opac.hs-kempten.de)  
            """)
        st.download_button("📄 PDF-Katalog herunterladen", data=b"", file_name="bibliothek_katalog.pdf")

    # Tab 4: Moodle
    with tabs[3]:
        st.subheader("💻 Moodle-Quicklink")
        user = st.text_input("Benutzername")
        pw = st.text_input("Passwort", type="password")
        if st.button("Login"):
            st.success("🔒 Simulierter Login erfolgreich")

      # Tab 5: Campus-Karte
    with tabs[4]:
        st.subheader("🗺️ Campus-Karte")

        # Interaktive Karte (auf Hochschule Kempten zentriert)
        df_map = pd.DataFrame({
            "lat": [47.72585],
            "lon": [10.31390]
        })
        st.map(df_map, zoom=18)

        st.markdown("---")
        st.markdown("### 🗺️ Gebäudeübersicht (Lageplan 2024)")
        st.image("https://www.hs-kempten.de/fileadmin/Bildpool/Lageplaene/Lageplan_Hochschule_Kempten_2024_DE.jpg", use_container_width=True)
        st.caption("Quelle: Hochschule Kempten")

        # PDF-Download 2024
        pdf_url = "https://www.hs-kempten.de/fileadmin/Bildpool/Lageplaene/Lageplan_Hochschule_Kempten_2024_DE.pdf"
        try:
            st.download_button(
                label="📥 Lageplan 2024 als PDF herunterladen",
                data=requests.get(pdf_url).content,
                file_name="Lageplan_Hochschule_Kempten_2024.pdf",
                mime="application/pdf"
            )
        except:
            st.warning("⚠️ Lageplan konnte nicht geladen werden.")


    st.markdown("---")
    st.info("🌟 Designed by dein Studi-Buddy 🚀")

